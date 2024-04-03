from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from selenium.webdriver.common.action_chains import ActionChains
from constants.generalConstants import *
from locators.generalLocators import *
import pytest
import openpyxl
import json
import time

class Test_Odev_4():
    def setup_method(self):
        self.driver = webdriver.Chrome()
        self.driver.maximize_window()
        self.driver.get(BASE_URL)

    def teardown_method(self):
        self.driver.quit()

    def waitForElementVisible(self,locator,timeout=10):
        return WebDriverWait(self.driver,timeout).until(ec.visibility_of_element_located(locator))

    def getData():
        return [("0","0"),("abc","123"),("tobeto",password_correct),(user_standart,"tobeto")]
    
    def getData_2():
        return [(user_standart,""),(user_locked_out,""),(user_problem,""),(user_perf_glitch,"")]
    
    def readInvalidDataFromExcel():
            excelFile = openpyxl.load_workbook("C:/Users/oguzh/OneDrive/Masaüstü/tobeto-test2/Tobeto Selenium Ödevleri/Ödev-3/data/invalidLogin.xlsx")
            sheet = excelFile["Sheet1"]
            rows = sheet.max_row #kaçıncı satıra kadar benim verim var
            data = []
            for i in range(2,rows+1):
                username = sheet.cell(i,1).value
                password = sheet.cell(i,2).value
                data.append((username,password))
            return data
    
    def readInvalidDataFromJson():
        with open("C:/Users/oguzh/OneDrive/Masaüstü/tobeto-test2/Tobeto Selenium Ödevleri/Ödev-3/data/invalidLogin.json", 'r') as json_file:
            data = json.load(json_file)
        return [(username, password) for username, password in zip(data["username"], data["password"])]

    # def test_empty_username(self):
    #     userNameInput = self.waitForElementVisible((By.ID,username_id))
    #     passwordInput = self.waitForElementVisible((By.ID,password_id))
    #     actions = ActionChains(self.driver)
    #     actions.send_keys_to_element(userNameInput,"")
    #     actions.send_keys_to_element(passwordInput,password_correct)
    #     actions.perform()
    #     loginButton = self.waitForElementVisible((By.ID,login_button_id))
    #     loginButton.click()
    #     errorMessage = WebDriverWait(self.driver, 10).until(ec.visibility_of_element_located((By.XPATH, "//*[@id='login_button_container']/div/form/div[3]/h3")))
    #     assert errorMessage.text == errorMessage_empty_username_text, errorMessage_not_as_expected


    def test_emptyusername(self):
        self.waitForElementVisible((By.ID, username_id))
        self.driver.find_element(By.ID, username_id).click()
        self.waitForElementVisible((By.ID, password_id))
        self.driver.find_element(By.ID, password_id).click()
        self.driver.find_element(By.ID, password_id).send_keys(password_correct)
        self.waitForElementVisible((By.ID, login_button_id))
        self.driver.find_element(By.ID, login_button_id).click()
        assert self.driver.find_element(By.XPATH, errorMessage_empty_username_xpath).text == errorMessage_empty_username_text

    # @pytest.mark.parametrize("username, password", getData_2())
    # def test_empty_password(self, username, password):
    #     userNameInput = self.waitForElementVisible((By.ID,username_id))
    #     passwordInput = self.waitForElementVisible((By.ID,password_id))
    #     userNameInput.send_keys(username)
    #     passwordInput.send_keys(password)
    #     loginButton = self.waitForElementVisible((By.ID,login_button_id))
    #     loginButton.click()
    #     errorMessage = WebDriverWait(self.driver, 10).until(ec.visibility_of_element_located((By.XPATH, "//*[@id='login_button_container']/div/form/div[3]/h3")))
    #     assert errorMessage.text == errorMessage_empty_password_text, errorMessage_not_as_expected
        
    def test_emptypassword(self):
        self.waitForElementVisible((By.ID, username_id))
        self.driver.find_element(By.ID, username_id).click()
        self.driver.find_element(By.ID, username_id).send_keys(user_standart)
        self.waitForElementVisible((By.ID, password_id))
        self.driver.find_element(By.ID, password_id).click()
        self.waitForElementVisible((By.ID, login_button_id))
        self.driver.find_element(By.ID, login_button_id).click()
        assert self.driver.find_element(By.XPATH, errorMessage_empty_password_xpath).text == errorMessage_empty_password_text


    # def test_locked_out_user(self):
    #     userNameInput = self.waitForElementVisible((By.ID,username_id))
    #     passwordInput = self.waitForElementVisible((By.ID,password_id))
    #     actions = ActionChains(self.driver)
    #     actions.send_keys_to_element(userNameInput,user_locked_out)
    #     actions.send_keys_to_element(passwordInput,password_correct)
    #     actions.perform()
    #     loginButton = self.waitForElementVisible((By.ID,login_button_id))
    #     loginButton.click()
    #     errorMessage = WebDriverWait(self.driver, 10).until(ec.visibility_of_element_located((By.XPATH, "//*[@id='login_button_container']/div/form/div[3]/h3")))
    #     assert errorMessage.text == errorMessage_lockout_user_text, errorMessage_not_as_expected
        
    def test_lockoutuser(self):
        self.waitForElementVisible((By.ID, username_id))
        self.driver.find_element(By.ID, username_id).click()
        self.driver.find_element(By.ID, username_id).send_keys(user_locked_out)
        self.waitForElementVisible((By.ID, password_id))
        self.driver.find_element(By.ID, password_id).click()
        self.driver.find_element(By.ID, password_id).send_keys(password_correct)
        self.waitForElementVisible((By.ID, login_button_id))
        self.driver.find_element(By.ID, login_button_id).click()
        assert self.driver.find_element(By.XPATH, errorMessage_lockout_user_xpath).text == errorMessage_lockout_user_text


    # def test_standard_user(self):
    #     userNameInput = self.waitForElementVisible((By.ID,username_id))
    #     passwordInput = self.waitForElementVisible((By.ID,password_id))
    #     actions = ActionChains(self.driver)
    #     actions.send_keys_to_element(userNameInput,user_standart)
    #     actions.send_keys_to_element(passwordInput,password_correct)
    #     actions.perform()
    #     loginButton = self.waitForElementVisible((By.ID,login_button_id))
    #     loginButton.click()
    #     WebDriverWait(self.driver, 10).until(ec.url_to_be("https://www.saucedemo.com/inventory.html"))
    #     listOfItems = WebDriverWait(self.driver, 10).until(ec.visibility_of_all_elements_located((By.CLASS_NAME, "inventory_item")))
    #     print(f"Saucedemo sitesinde şu an {len(listOfItems)} adet ürün bulunmaktadır.")
        
    def test_standarduser(self):
        self.waitForElementVisible((By.ID, username_id))
        self.driver.find_element(By.ID, username_id).click()
        self.driver.find_element(By.ID, username_id).send_keys(user_standart)
        self.waitForElementVisible((By.ID, password_id))
        self.driver.find_element(By.ID, password_id).click()
        self.driver.find_element(By.ID, password_id).send_keys(password_correct)
        self.waitForElementVisible((By.ID, login_button_id))
        self.driver.find_element(By.ID, login_button_id).click()
        self.vars = {}  # Seleniumdan gelen kodda bu kısım yok fakat burası olmadan hata veriyor.
        self.vars["the_url"] = self.driver.execute_script("return document.URL")
        assert(self.vars["the_url"] == "https://www.saucedemo.com/inventory.html")
        self.vars["elementCount"] = len(self.driver.find_elements(By.XPATH, "//div[2]/div/div/div/div"))
        assert(self.vars["elementCount"] == 6)


    # @pytest.mark.parametrize("username, password", readInvalidDataFromExcel())
    # def test_invalid_login(self, username, password):
    #     userNameInput = self.waitForElementVisible((By.ID,username_id))
    #     passwordInput = self.waitForElementVisible((By.ID,password_id))
    #     userNameInput.send_keys(username)
    #     passwordInput.send_keys(password)
    #     loginButton = self.waitForElementVisible((By.ID,login_button_id))
    #     loginButton.click()
    #     errorMessage = self.waitForElementVisible((By.XPATH, errorMessage_xpath))
    #     assert errorMessage.text == errorMessage_text, errorMessage_not_as_expected


    # @pytest.mark.parametrize("username, password", readInvalidDataFromJson())
    # def test_invalid_login(self, username, password):
    #     userNameInput = self.waitForElementVisible((By.ID,username_id))
    #     passwordInput = self.waitForElementVisible((By.ID,password_id))
    #     userNameInput.send_keys(username)
    #     passwordInput.send_keys(password)
    #     loginButton = self.waitForElementVisible((By.ID,login_button_id))
    #     loginButton.click()
    #     errorMessage = self.waitForElementVisible((By.XPATH, errorMessage_xpath))
    #     assert errorMessage.text == errorMessage_text, errorMessage_not_as_expected
        
    def test_invalidlogin(self):
        self.waitForElementVisible((By.ID, username_id))
        self.driver.find_element(By.ID, username_id).click()
        self.driver.find_element(By.ID, username_id).send_keys("0")
        self.waitForElementVisible((By.ID, password_id))
        self.driver.find_element(By.ID, password_id).click()
        self.driver.find_element(By.ID, password_id).send_keys("0")
        self.waitForElementVisible((By.ID, login_button_id))
        self.driver.find_element(By.ID, login_button_id).click()
        assert self.driver.find_element(By.XPATH, errorMessage_invalid_login_xpath).text == errorMessage_invalid_login_text


    # def test_add_to_cart(self):
    #     userNameInput = self.waitForElementVisible((By.ID,username_id))
    #     passwordInput =self.waitForElementVisible((By.ID,password_id))
    #     actions = ActionChains(self.driver)
    #     actions.send_keys_to_element(userNameInput,user_standart)
    #     actions.send_keys_to_element(passwordInput,password_correct)
    #     actions.perform()
    #     loginButton = self.waitForElementVisible((By.ID,login_button_id))
    #     loginButton.click()
    #     addToCart = self.waitForElementVisible((By.ID,add_backpack_id))
    #     addToCart.click()
    #     shopping_cart_badge = self.waitForElementVisible((By.XPATH,"//*[@id='shopping_cart_container']/a/span"))
    #     assert shopping_cart_badge.text == "1"
        
    def test_addtocart(self):
        self.waitForElementVisible((By.ID, username_id))
        self.driver.find_element(By.ID, username_id).click()
        self.driver.find_element(By.ID, username_id).send_keys(user_standart)
        self.waitForElementVisible((By.ID, password_id))
        self.driver.find_element(By.ID, password_id).click()
        self.driver.find_element(By.ID, password_id).send_keys(password_correct)
        self.waitForElementVisible((By.ID, login_button_id))
        self.driver.find_element(By.ID, login_button_id).click()
        self.waitForElementVisible((By.ID, add_backpack_id))
        self.driver.find_element(By.ID, add_backpack_id).click()
        self.waitForElementVisible((By.XPATH, shopping_cart_container_xpath))
        assert self.driver.find_element(By.XPATH, shopping_cart_container_xpath).text == "1"

    # def test_remove(self):
    #     userNameInput = self.waitForElementVisible((By.ID,username_id))
    #     passwordInput =self.waitForElementVisible((By.ID,password_id))
    #     actions = ActionChains(self.driver)
    #     actions.send_keys_to_element(userNameInput,user_standart)
    #     actions.send_keys_to_element(passwordInput,password_correct)
    #     actions.perform()
    #     loginButton = self.waitForElementVisible((By.ID,login_button_id))
    #     loginButton.click()
    #     addToCart = self.waitForElementVisible((By.ID,add_backpack_id))
    #     addToCart.click()
    #     removeButton = self.waitForElementVisible((By.ID,remove_backpack_id))
    #     assert removeButton.text == "Remove"
    #     removeButton.click()
    #     addToCart = self.waitForElementVisible((By.ID,add_backpack_id))
    #     assert addToCart.text == "Add to cart"
        
    def test_remove(self):
        self.waitForElementVisible((By.ID, username_id))
        self.driver.find_element(By.ID, username_id).click()
        self.driver.find_element(By.ID, username_id).send_keys(user_standart)
        self.waitForElementVisible((By.ID, password_id))
        self.driver.find_element(By.ID, password_id).click()
        self.driver.find_element(By.ID, password_id).send_keys(password_correct)
        self.waitForElementVisible((By.ID, login_button_id))
        self.driver.find_element(By.ID, login_button_id).click()
        self.waitForElementVisible((By.ID, add_backpack_id))
        self.driver.find_element(By.ID, add_backpack_id).click()
        self.waitForElementVisible((By.ID, remove_backpack_id))
        assert self.driver.find_element(By.ID, remove_backpack_id).text == remove_backpack_text
        self.driver.find_element(By.ID, remove_backpack_id).click()
        self.waitForElementVisible((By.ID, add_backpack_id))
        assert self.driver.find_element(By.ID, add_backpack_id).text == add_backpack_text


    # def test_checkout(self):
    #     userNameInput = self.waitForElementVisible((By.ID,username_id))
    #     passwordInput =self.waitForElementVisible((By.ID,password_id))
    #     actions = ActionChains(self.driver)
    #     actions.send_keys_to_element(userNameInput,user_standart)
    #     actions.send_keys_to_element(passwordInput,password_correct)
    #     actions.perform()
    #     loginButton = self.waitForElementVisible((By.ID,login_button_id))
    #     loginButton.click()
    #     addToCart = self.waitForElementVisible((By.ID,add_backpack_id))
    #     addToCart.click()
    #     shoppingCartButton = self.waitForElementVisible((By.CLASS_NAME,"shopping_cart_link"))
    #     shoppingCartButton.click()
    #     title_1 = self.waitForElementVisible((By.CLASS_NAME, title_name))
    #     assert title_1.text == "Your Cart"
    #     checkoutButton = self.waitForElementVisible((By.ID,checkout_button_id))
    #     checkoutButton.click()
    #     title_2 = self.waitForElementVisible((By.CLASS_NAME, title_name))
    #     assert title_2.text == "Checkout: Your Information"
    #     firstNameInput = self.waitForElementVisible((By.ID,firstname_id))
    #     lastNameInput =self.waitForElementVisible((By.ID,lastname_id))
    #     postalCodeInput =self.waitForElementVisible((By.ID,postalcode_id))
    #     actions = ActionChains(self.driver)
    #     actions.send_keys_to_element(firstNameInput,"1")
    #     actions.send_keys_to_element(lastNameInput,"1")
    #     actions.send_keys_to_element(postalCodeInput,"1")
    #     actions.perform()
    #     continueButton = self.waitForElementVisible((By.ID,continue_button_id))
    #     continueButton.click()
    #     title_3 = self.waitForElementVisible((By.CLASS_NAME, title_name))
    #     assert title_3.text == "Checkout: Overview"
    #     finishButton = self.waitForElementVisible((By.ID,finish_button_id))
    #     finishButton.click()
    #     thanksMessage = self.waitForElementVisible((By.CLASS_NAME, "complete-header"))
    #     assert thanksMessage.text == "Thank you for your order!"
        
    def test_checkout(self):
        self.waitForElementVisible((By.ID, username_id))
        self.driver.find_element(By.ID, username_id).click()
        self.driver.find_element(By.ID, username_id).send_keys(user_standart)
        self.waitForElementVisible((By.ID, password_id))
        self.driver.find_element(By.ID, password_id).click()
        self.driver.find_element(By.ID, password_id).send_keys(password_correct)
        self.waitForElementVisible((By.ID, login_button_id))
        self.driver.find_element(By.ID, login_button_id).click()
        self.waitForElementVisible((By.ID, add_backpack_id))
        self.driver.find_element(By.ID, add_backpack_id).click()
        self.waitForElementVisible((By.CSS_SELECTOR, shopping_cart_css))
        self.driver.find_element(By.CSS_SELECTOR, shopping_cart_css).click()
        self.waitForElementVisible((By.CSS_SELECTOR, title_css))
        assert self.driver.find_element(By.CSS_SELECTOR, title_css).text == title_1_text
        self.waitForElementVisible((By.ID, checkout_button_id))
        self.driver.find_element(By.ID, checkout_button_id).click()
        self.waitForElementVisible((By.CSS_SELECTOR, title_css))
        assert self.driver.find_element(By.CSS_SELECTOR, title_css).text == title_2_text
        self.waitForElementVisible((By.ID, firstname_id))
        self.driver.find_element(By.ID, firstname_id).click()
        self.driver.find_element(By.ID, firstname_id).send_keys("1")
        self.waitForElementVisible((By.ID, lastname_id))
        self.driver.find_element(By.ID, lastname_id).click()
        self.driver.find_element(By.ID, lastname_id).send_keys("1")
        self.waitForElementVisible((By.ID, postalcode_id))
        self.driver.find_element(By.ID, postalcode_id).click()
        self.driver.find_element(By.ID, postalcode_id).send_keys("1")
        self.waitForElementVisible((By.ID, continue_button_id))
        self.driver.find_element(By.ID, continue_button_id).click()
        self.waitForElementVisible((By.CSS_SELECTOR, title_css))
        assert self.driver.find_element(By.CSS_SELECTOR, title_css).text == title_3_text
        self.waitForElementVisible((By.ID, finish_button_id))
        self.driver.find_element(By.ID, finish_button_id).click()
        self.waitForElementVisible((By.CSS_SELECTOR, complete_header_css))
        assert self.driver.find_element(By.CSS_SELECTOR, complete_header_css).text == complete_header_text

